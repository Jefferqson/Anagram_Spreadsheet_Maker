from openpyxl import Workbook


def spreadsheet_generator(arranged_pairs):
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Anagram"
    ws["B1"] = "Words"
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row + 1) and arranged_pairs:
        ws.append(row)

    wb.save("AnagramDict.xlsx")
